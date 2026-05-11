import asyncio
import os
import re
from datetime import datetime, timedelta, timezone
from io import BytesIO
from aiogram import Bot, Dispatcher, F, types
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery, BufferedInputFile
from supabase import create_client, Client
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import aiohttp
from aiohttp import web

# === КОНФИГ ===
BOT_TOKEN = os.getenv("BOT_TOKEN")
OWNER_ID = int(os.getenv("OWNER_ID", 6810564564))
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

bot = Bot(token=BOT_TOKEN)
storage = MemoryStorage()
dp = Dispatcher(storage=storage)
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# === КЛАВИАТУРЫ ===
main_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Новый трек")],
        [KeyboardButton(text="📋 Мои треки"), KeyboardButton(text="📊 Мои треки (Excel)")],
        [KeyboardButton(text="Редактировать профиль"), KeyboardButton(text="Удалить трек")],
        [KeyboardButton(text="💰 Конвертер валют"), KeyboardButton(text="🚚 Калькулятор доставки")],
        [KeyboardButton(text="✅ Завершить и отправить"), KeyboardButton(text="🗑 Удалить все треки")]
    ],
    resize_keyboard=True
)

owner_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Новый трек")],
        [KeyboardButton(text="📋 Мои треки"), KeyboardButton(text="📊 Мои треки (Excel)")],
        [KeyboardButton(text="Редактировать профиль"), KeyboardButton(text="Удалить трек")],
        [KeyboardButton(text="💰 Конвертер валют"), KeyboardButton(text="🚚 Калькулятор доставки")],
        [KeyboardButton(text="✅ Завершить и отправить"), KeyboardButton(text="🗑 Удалить все треки")],
        [KeyboardButton(text="📢 Сделать рассылку")]
    ],
    resize_keyboard=True
)

cancel_keyboard = ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text="Отмена")]], resize_keyboard=True)
group_keyboard = InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="📦 Загрузить трек", url="https://t.me/little_Bro_track_bot")]])

# === FSM ===
class ProfileForm(StatesGroup):
    waiting_for_fullname = State()
    waiting_for_phone = State()

class TrackForm(StatesGroup):
    waiting_for_track = State()
    waiting_for_product = State()
    waiting_for_price_cny = State()
    waiting_for_quantity_type = State()
    waiting_for_quantity = State()

class DeleteTrackForm(StatesGroup):
    waiting_for_track_ids = State()

class CurrencyForm(StatesGroup):
    waiting_for_amount = State()

class CalcForm(StatesGroup):
    waiting_for_city = State()
    waiting_for_weight = State()

class BroadcastForm(StatesGroup):
    waiting_for_start_date = State()
    waiting_for_end_date = State()
    waiting_for_text = State()

class ConfirmDeleteAllForm(StatesGroup):
    waiting_for_confirmation = State()

# === БАЗА ДАННЫХ ===
def get_msk_time():
    return datetime.now(timezone.utc) + timedelta(hours=3)

def get_user_profile(user_id):
    try:
        res = supabase.table("users").select("full_name, phone").eq("user_id", user_id).execute()
        if res.data:
            return res.data[0]["full_name"], res.data[0]["phone"]
        return None
    except:
        return None

def save_user_profile(user_id, username, full_name, phone):
    supabase.table("users").upsert({
        "user_id": user_id,
        "username": username,
        "full_name": full_name,
        "phone": phone,
        "created_at": get_msk_time().isoformat()
    }).execute()

def add_track(user_id, track, product, price_cny, price_usd, price_byn, qty_type, quantity):
    supabase.table("tracks").insert({
        "user_id": user_id,
        "track_number": track,
        "product_name": product,
        "price_cny": price_cny,
        "price_usd": price_usd,
        "price_byn": price_byn,
        "quantity": quantity,
        "quantity_type": qty_type,
        "created_at": get_msk_time().isoformat()
    }).execute()

def get_user_tracks(user_id):
    try:
        res = supabase.table("tracks").select("*").eq("user_id", user_id).order("created_at", desc=True).execute()
        return res.data
    except:
        return []

def delete_track_by_id(track_id, user_id):
    supabase.table("tracks").delete().eq("id", track_id).eq("user_id", user_id).execute()

def delete_all_tracks(user_id):
    supabase.table("tracks").delete().eq("user_id", user_id).execute()

def get_total_sum_cny(user_id):
    tracks = get_user_tracks(user_id)
    return round(sum(t["price_cny"] * t["quantity"] for t in tracks), 2) if tracks else 0

def get_total_sum_usd(user_id):
    tracks = get_user_tracks(user_id)
    return round(sum((t.get("price_usd") or 0) * t["quantity"] for t in tracks), 2) if tracks else 0

def get_total_sum_byn(user_id):
    tracks = get_user_tracks(user_id)
    return round(sum((t.get("price_byn") or 0) * t["quantity"] for t in tracks), 2) if tracks else 0

def get_total_quantity(user_id):
    tracks = get_user_tracks(user_id)
    return sum(t["quantity"] for t in tracks) if tracks else 0

# === ПОЛУЧЕНИЕ КУРСОВ ОТ НАЦБАНКА БЕЛАРУСИ (с учётом масштаба) ===
async def get_rates_from_nbrb():
    """
    Возвращает (usd_to_byn, cny_to_byn) - официальные курсы НБРБ.
    """
    usd_to_byn = None
    cny_to_byn = None
    fallback_usd = 3.2
    fallback_cny = 0.45

    try:
        async with aiohttp.ClientSession() as session:
            # Курс USD (обычно масштаб 1)
            async with session.get("https://api.nbrb.by/exrates/rates/USD?parammode=2", timeout=10) as resp:
                if resp.status == 200:
                    data = await resp.json()
                    usd_to_byn = data.get("Cur_OfficialRate")
                    print(f"Курс USD/BYN от Нацбанка: {usd_to_byn}")

            # Курс CNY (масштаб может быть 10)
            async with session.get("https://api.nbrb.by/exrates/rates/CNY?parammode=2", timeout=10) as resp:
                if resp.status == 200:
                    data = await resp.json()
                    official_rate = data.get("Cur_OfficialRate")
                    scale = data.get("Cur_Scale", 1)
                    cny_to_byn = official_rate / scale
                    print(f"Курс CNY/BYN от Нацбанка (за 1 CNY): {cny_to_byn} (официальный {official_rate} за {scale} CNY)")
    except Exception as e:
        print(f"Ошибка получения курсов от Нацбанка: {e}")

    if usd_to_byn is None:
        usd_to_byn = fallback_usd
        print(f"Используем резервный USD/BYN: {usd_to_byn}")
    if cny_to_byn is None:
        cny_to_byn = fallback_cny
        print(f"Используем резервный CNY/BYN: {cny_to_byn}")

    return usd_to_byn, cny_to_byn

async def get_cny_to_usd_rate():
    usd_to_byn, cny_to_byn = await get_rates_from_nbrb()
    return cny_to_byn / usd_to_byn if usd_to_byn else 0.14

async def get_usd_to_byn_rate():
    usd_to_byn, _ = await get_rates_from_nbrb()
    return usd_to_byn

async def get_cny_to_byn_rate():
    _, cny_to_byn = await get_rates_from_nbrb()
    return cny_to_byn

# === EXCEL ===
def create_excel(tracks, full_name, phone, user_id):
    wb = Workbook()
    ws = wb.active
    ws.title = "Треки"
    headers = ["№", "Трек-номер", "Товар", "Цена (CNY)", "Цена (USD)", "Цена (BYN)", "Кол-во", "Ед. изм.", "Дата"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for i, t in enumerate(tracks, 2):
        ws.cell(row=i, column=1, value=i-1)
        ws.cell(row=i, column=2, value=t["track_number"])
        ws.cell(row=i, column=3, value=t["product_name"])
        ws.cell(row=i, column=4, value=float(t["price_cny"]))
        ws.cell(row=i, column=5, value=float(t.get("price_usd") or 0))
        ws.cell(row=i, column=6, value=float(t.get("price_byn") or 0))
        ws.cell(row=i, column=7, value=int(t["quantity"]))
        ws.cell(row=i, column=8, value=t["quantity_type"])
        dt = datetime.fromisoformat(t['created_at'])
        ws.cell(row=i, column=9, value=dt.strftime("%Y-%m-%d %H:%M:%S"))
    last = len(tracks) + 2
    ws.cell(row=last, column=1, value="Всего треков:"); ws.cell(row=last, column=2, value=len(tracks))
    ws.cell(row=last+1, column=1, value="Общее количество:"); ws.cell(row=last+1, column=2, value=get_total_quantity(user_id))
    ws.cell(row=last+2, column=7, value="ИТОГО (CNY):"); ws.cell(row=last+2, column=8, value=f"{get_total_sum_cny(user_id):.2f}")
    ws.cell(row=last+3, column=7, value="ИТОГО (USD):"); ws.cell(row=last+3, column=8, value=f"{get_total_sum_usd(user_id):.2f}")
    ws.cell(row=last+4, column=7, value="ИТОГО (BYN):"); ws.cell(row=last+4, column=8, value=f"{get_total_sum_byn(user_id):.2f}")
    ws.cell(row=last+6, column=1, value=f"ФИО: {full_name}")
    ws.cell(row=last+7, column=1, value=f"Телефон: {phone}")
    ws.cell(row=last+8, column=1, value=f"ID: {user_id}")
    for col in range(1, 10):
        ws.column_dimensions[chr(64+col)].width = 18
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# === ХЕНДЛЕРЫ ===
@dp.message(Command("start"))
async def cmd_start(message: types.Message, state: FSMContext):
    if message.chat.type in ["group", "supergroup"]:
        await message.answer("📦 **Загрузка треков**\n\nНажми на кнопку:", reply_markup=group_keyboard, parse_mode="Markdown")
        return
    profile = get_user_profile(message.from_user.id)
    keyboard = owner_keyboard if message.from_user.id == OWNER_ID else main_keyboard
    if profile:
        await message.answer(f"✅ С возвращением!\n\nФИО: {profile[0]}\nТелефон: {profile[1]}", reply_markup=keyboard)
    else:
        await state.set_state(ProfileForm.waiting_for_fullname)
        await message.answer("Введи твоё ФИО:", reply_markup=cancel_keyboard)

@dp.message(ProfileForm.waiting_for_fullname)
async def process_fullname(message: types.Message, state: FSMContext):
    if message.text == "Отмена":
        await cancel_action(message, state)
        return
    await state.update_data(fullname=message.text)
    await state.set_state(ProfileForm.waiting_for_phone)
    await message.answer("Введи номер телефона:", reply_markup=cancel_keyboard)

@dp.message(ProfileForm.waiting_for_phone)
async def process_phone(message: types.Message, state: FSMContext):
    if message.text == "Отмена":
        await cancel_action(message, state)
        return
    data = await state.get_data()
    full_name = data.get("fullname")
    if not full_name:
        await message.answer("Ошибка. Начните заново с /start")
        await state.clear()
        return
    save_user_profile(message.from_user.id, message.from_user.username or "нет", full_name, message.text.strip())
    await state.clear()
    keyboard = owner_keyboard if message.from_user.id == OWNER_ID else main_keyboard
    await message.answer("✅ Профиль сохранён! Теперь можно добавлять треки.", reply_markup=keyboard)

@dp.message(F.text == "Редактировать профиль")
async def edit_profile(message: types.Message, state: FSMContext):
    await state.set_state(ProfileForm.waiting_for_fullname)
    await message.answer("Введи новое ФИО:", reply_markup=cancel_keyboard)

@dp.message(F.text == "Отмена")
async def cancel_action(message: types.Message, state: FSMContext):
    await state.clear()
    keyboard = owner_keyboard if message.from_user.id == OWNER_ID else main_keyboard
    await message.answer("Действие отменено", reply_markup=keyboard)

@dp.message(F.text == "Новый трек")
async def new_track(message: types.Message, state: FSMContext):
    if not get_user_profile(message.from_user.id):
        await state.set_state(ProfileForm.waiting_for_fullname)
        await message.answer("Сначала заполни профиль. Введи ФИО:", reply_markup=cancel_keyboard)
        return
    await state.set_state(TrackForm.waiting_for_track)
    await message.answer("Введи трек-номер:", reply_markup=cancel_keyboard)

@dp.message(TrackForm.waiting_for_track)
async def process_track(message: types.Message, state: FSMContext):
    if message.text == "Отмена":
        await cancel_action(message, state)
        return
    await state.update_data(track=message.text)
    await state.set_state(TrackForm.waiting_for_product)
    await message.answer("Введи наименование товара:", reply_markup=cancel_keyboard)

@dp.message(TrackForm.waiting_for_product)
async def process_product(message: types.Message, state: FSMContext):
    if message.text == "Отмена":
        await cancel_action(message, state)
        return
    await state.update_data(product=message.text)
    await state.set_state(TrackForm.waiting_for_price_cny)
    await message.answer("Введи цену в **юанях (CNY)**:\n(бот пересчитает в USD и BYN по курсу Нацбанка)", reply_markup=cancel_keyboard, parse_mode="Markdown")

@dp.message(TrackForm.waiting_for_price_cny)
async def process_price_cny(message: types.Message, state: FSMContext):
    if message.text == "Отмена":
        await cancel_action(message, state)
        return
    match = re.search(r"(\d+(?:[.,]\d+)?)", message.text)
    if not match:
        await message.answer("Введи число!", reply_markup=cancel_keyboard)
        return
    price_cny = float(match.group(1).replace(",", "."))
    usd_to_byn, cny_to_byn = await get_rates_from_nbrb()
    price_byn = round(price_cny * cny_to_byn, 2)
    price_usd = round(price_byn / usd_to_byn, 2)
    await state.update_data(price_cny=price_cny, price_usd=price_usd, price_byn=price_byn)
    await state.set_state(TrackForm.waiting_for_quantity_type)
    await message.answer(f"💰 Цена: {price_cny:.2f} CNY = {price_usd:.2f} USD = {price_byn:.2f} BYN\n\nВыбери единицу измерения:", reply_markup=ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text="шт"), KeyboardButton(text="пара")]], resize_keyboard=True))

@dp.message(TrackForm.waiting_for_quantity_type)
async def process_quantity_type(message: types.Message, state: FSMContext):
    if message.text not in ["шт", "пара"]:
        await message.answer("Выбери кнопку: шт или пара")
        return
    await state.update_data(qtype=message.text)
    await state.set_state(TrackForm.waiting_for_quantity)
    await message.answer(f"Введи количество в {message.text}:", reply_markup=cancel_keyboard)

@dp.message(TrackForm.waiting_for_quantity)
async def process_quantity(message: types.Message, state: FSMContext):
    if message.text == "Отмена":
        await cancel_action(message, state)
        return
    try:
        qty = int(message.text)
    except:
        await message.answer("Введи целое число!")
        return
    data = await state.get_data()
    add_track(message.from_user.id, data["track"], data["product"], data["price_cny"], data["price_usd"], data["price_byn"], data["qtype"], qty)
    await state.clear()
    keyboard = owner_keyboard if message.from_user.id == OWNER_ID else main_keyboard
    await message.answer("✅ Трек добавлен!", reply_markup=keyboard)

@dp.message(F.text == "📋 Мои треки")
async def my_tracks(message: types.Message):
    tracks = get_user_tracks(message.from_user.id)
    if not tracks:
        await message.answer("Нет треков.", reply_markup=main_keyboard)
        return
    text = "📦 ТВОИ ТРЕКИ:\n\n"
    for i, t in enumerate(tracks, 1):
        dt = datetime.fromisoformat(t['created_at'])
        usd = t.get('price_usd', 0) or 0
        byn = t.get('price_byn', 0) or 0
        text += f"{i}. {t['track_number']}\n   {t['product_name']}\n   Цена: {t['price_cny']:.2f} CNY = {usd:.2f} USD = {byn:.2f} BYN\n   Кол-во: {t['quantity']} {t['quantity_type']}\n   Дата: {dt.strftime('%Y-%m-%d %H:%M:%S')}\n\n"
    text += f"💰 **Итого: {get_total_sum_cny(message.from_user.id):.2f} CNY ≈ {get_total_sum_usd(message.from_user.id):.2f} USD ≈ {get_total_sum_byn(message.from_user.id):.2f} BYN**"
    await message.answer(text, reply_markup=main_keyboard, parse_mode="Markdown")

@dp.message(F.text == "Удалить трек")
async def delete_track_start(message: types.Message, state: FSMContext):
    tracks = get_user_tracks(message.from_user.id)
    if not tracks:
        await message.answer("Нет треков для удаления.", reply_markup=main_keyboard)
        return
    text = "🗑 Введи **номера** треков для удаления (через запятую, например: 1,3,5):\n"
    for i, t in enumerate(tracks, 1):
        text += f"{i}. {t['track_number']} - {t['product_name']}\n"
    await state.set_state(DeleteTrackForm.waiting_for_track_ids)
    await state.update_data(tracks=tracks)
    await message.answer(text, reply_markup=cancel_keyboard)

@dp.message(DeleteTrackForm.waiting_for_track_ids)
async def process_delete_tracks(message: types.Message, state: FSMContext):
    if message.text == "Отмена":
        await cancel_action(message, state)
        return
    data = await state.get_data()
    tracks = data["tracks"]
    parts = message.text.replace(" ", "").split(",")
    deleted = 0
    for part in parts:
        try:
            idx = int(part) - 1
            if 0 <= idx < len(tracks):
                delete_track_by_id(tracks[idx]["id"], message.from_user.id)
                deleted += 1
        except:
            continue
    keyboard = owner_keyboard if message.from_user.id == OWNER_ID else main_keyboard
    await message.answer(f"✅ Удалено треков: {deleted}", reply_markup=keyboard)
    await state.clear()

@dp.message(F.text == "🗑 Удалить все треки")
async def delete_all_tracks_start(message: types.Message, state: FSMContext):
    if not get_user_tracks(message.from_user.id):
        await message.answer("Нет треков.", reply_markup=main_keyboard)
        return
    await state.set_state(ConfirmDeleteAllForm.waiting_for_confirmation)
    await message.answer("⚠️ Удалить ВСЕ треки? Напишите `ДА` заглавными", reply_markup=cancel_keyboard, parse_mode="Markdown")

@dp.message(ConfirmDeleteAllForm.waiting_for_confirmation)
async def confirm_delete_all(message: types.Message, state: FSMContext):
    if message.text == "Отмена":
        await cancel_action(message, state)
        return
    if message.text.strip() == "ДА":
        delete_all_tracks(message.from_user.id)
        keyboard = owner_keyboard if message.from_user.id == OWNER_ID else main_keyboard
        await message.answer("✅ Все треки удалены.", reply_markup=keyboard)
    else:
        await message.answer("❌ Отменено.", reply_markup=main_keyboard)
    await state.clear()

@dp.message(F.text == "📊 Мои треки (Excel)")
async def export_excel(message: types.Message):
    tracks = get_user_tracks(message.from_user.id)
    if not tracks:
        await message.answer("Нет треков.", reply_markup=main_keyboard)
        return
    prof = get_user_profile(message.from_user.id)
    if not prof:
        await message.answer("Профиль не найден. /start")
        return
    try:
        excel_file = create_excel(tracks, prof[0], prof[1], message.from_user.id)
        await message.answer_document(BufferedInputFile(excel_file.getvalue(), filename=f"tracks_{message.from_user.id}.xlsx"), caption="📊 Ваши треки")
    except Exception as e:
        await message.answer(f"Ошибка Excel: {e}")

@dp.message(F.text == "🚚 Калькулятор доставки")
async def calc_button(message: types.Message, state: FSMContext):
    kb = ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text="Минск"), KeyboardButton(text="Лида")]], resize_keyboard=True)
    await state.set_state(CalcForm.waiting_for_city)
    await message.answer("Выберите город:", reply_markup=kb)

@dp.message(CalcForm.waiting_for_city)
async def calc_weight(message: types.Message, state: FSMContext):
    if message.text not in ["Минск", "Лида"]:
        await message.answer("Кнопкой!")
        return
    await state.update_data(city=message.text)
    await state.set_state(CalcForm.waiting_for_weight)
    await message.answer("Введите вес (кг):", reply_markup=cancel_keyboard)

@dp.message(CalcForm.waiting_for_weight)
async def calc_result(message: types.Message, state: FSMContext):
    try:
        weight = float(message.text.replace(',', '.'))
    except:
        await message.answer("Введите число!")
        return
    data = await state.get_data()
    city = data.get("city")
    
    # Все параметры доставки из переменных окружения (можно менять в панели Render)
    delivery_rate = float(os.getenv("DELIVERY_RATE", "12.0"))           # руб/кг базовая доставка
    handling_rate = float(os.getenv("HANDLING_RATE", "1.6"))           # руб/кг обработка
    lida_extra_rate = float(os.getenv("LIDA_EXTRA_RATE", "0.8"))       # руб/кг доплата за Лиду
    fixed_fee = float(os.getenv("FIXED_FEE", "10.0"))                  # руб фиксированный сбор
    # Дополнительно: доставка Москва-Минск (по умолчанию 0, но можно изменить)
    moscow_delivery_rate = float(os.getenv("DELIVERY_MOSCOW_MINSK", "0.0"))
    
    # Если вдруг есть старая переменная DELIVERY_MINSK_LIDA, используем её как lida_extra_rate
    if os.getenv("DELIVERY_MINSK_LIDA"):
        lida_extra_rate = float(os.getenv("DELIVERY_MINSK_LIDA"))
    
    # Расчёт
    cost = weight * delivery_rate + weight * handling_rate + (weight * lida_extra_rate if city == "Лида" else 0) + fixed_fee
    # Если нужна доставка из Москвы в Минск, можно добавить условие, но сейчас город только Минск и Лида
    # Для простоты пока оставляем как есть.
    
    await message.answer(f"🚚 Доставка до {city}: {cost:.2f} руб.", reply_markup=main_keyboard)
    await state.clear()

@dp.message(F.text == "💰 Конвертер валют")
async def currency_button(message: types.Message, state: FSMContext):
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="CNY → USD → BYN", callback_data="cny_all")],
        [InlineKeyboardButton(text="USD → BYN", callback_data="usd2byn")],
        [InlineKeyboardButton(text="BYN → USD", callback_data="byn2usd")],
        [InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_curr")]
    ])
    await message.answer("Выберите операцию:", reply_markup=kb)

@dp.callback_query(lambda c: c.data in ["cny_all", "usd2byn", "byn2usd", "cancel_curr"])
async def currency_callback(call: CallbackQuery, state: FSMContext):
    await call.answer()
    if call.data == "cancel_curr":
        await call.message.edit_text("Отменено")
        return
    await state.update_data(conv_action=call.data)
    await state.set_state(CurrencyForm.waiting_for_amount)
    await call.message.edit_text("Введите сумму:")

@dp.message(CurrencyForm.waiting_for_amount)
async def process_currency_amount(message: types.Message, state: FSMContext):
    try:
        amount = float(message.text.replace(',', '.'))
    except:
        await message.answer("Введите число!")
        return
    data = await state.get_data()
    action = data.get("conv_action")
    usd_to_byn, cny_to_byn = await get_rates_from_nbrb()
    cny_to_usd = cny_to_byn / usd_to_byn if usd_to_byn else 0.14
    if action == "cny_all":
        byn = amount * cny_to_byn
        usd = amount * cny_to_usd
        await message.answer(f"{amount:.2f} CNY = {usd:.2f} USD = {byn:.2f} BYN")
    elif action == "usd2byn":
        byn = amount * usd_to_byn
        await message.answer(f"{amount:.2f} USD = {byn:.2f} BYN")
    elif action == "byn2usd":
        usd = amount / usd_to_byn
        await message.answer(f"{amount:.2f} BYN = {usd:.2f} USD")
    else:
        await message.answer("Ошибка")
    await state.clear()

@dp.message(F.text == "✅ Завершить и отправить")
async def finish_and_send(message: types.Message):
    user_id = message.from_user.id
    tracks = get_user_tracks(user_id)
    if not tracks:
        await message.answer("Нет треков.")
        return
    prof = get_user_profile(user_id)
    if not prof:
        await message.answer("Профиль не найден.")
        return
    full_name, phone = prof
    text = f"📦 ТРЕКИ ПОЛЬЗОВАТЕЛЯ\n👤 {message.from_user.full_name} (@{message.from_user.username or 'нет'})\n📝 {full_name}\n📞 {phone}\n\n"
    for i, t in enumerate(tracks, 1):
        dt = datetime.fromisoformat(t['created_at'])
        usd = t.get('price_usd', 0) or 0
        byn = t.get('price_byn', 0) or 0
        text += f"{i}. {t['track_number']} – {t['product_name']}\n   Цена: {t['price_cny']:.2f} CNY ≈ {usd:.2f} USD ≈ {byn:.2f} BYN\n   Кол-во: {t['quantity']} {t['quantity_type']} ({dt.strftime('%Y-%m-%d %H:%M:%S')})\n\n"
    text += f"💰 Итого: {get_total_sum_cny(user_id):.2f} CNY ≈ {get_total_sum_usd(user_id):.2f} USD ≈ {get_total_sum_byn(user_id):.2f} BYN"
    excel_file = create_excel(tracks, full_name, phone, user_id)
    try:
        await bot.send_message(OWNER_ID, text)
        await bot.send_document(OWNER_ID, BufferedInputFile(excel_file.getvalue(), filename=f"tracks_{user_id}.xlsx"), caption=f"Excel от {full_name}")
        await message.answer("✅ Отправлено владельцу!")
    except Exception as e:
        await message.answer(f"Ошибка: {e}")

@dp.message(F.text == "📢 Сделать рассылку")
async def broadcast_start(message: types.Message, state: FSMContext):
    if message.from_user.id != OWNER_ID:
        await message.answer("⛔ Только для владельца.")
        return
    await state.set_state(BroadcastForm.waiting_for_start_date)
    await message.answer("Введите начальную дату (ГГГГ-ММ-ДД):", reply_markup=cancel_keyboard)

@dp.message(BroadcastForm.waiting_for_start_date)
async def broadcast_start_date(message: types.Message, state: FSMContext):
    if message.text == "Отмена":
        await cancel_action(message, state)
        return
    try:
        start_date = datetime.strptime(message.text.strip(), "%Y-%m-%d")
        await state.update_data(start_date=start_date)
        await state.set_state(BroadcastForm.waiting_for_end_date)
        await message.answer("Введите конечную дату (ГГГГ-ММ-ДД):")
    except:
        await message.answer("Неверный формат. Используйте ГГГГ-ММ-ДД")

@dp.message(BroadcastForm.waiting_for_end_date)
async def broadcast_end_date(message: types.Message, state: FSMContext):
    if message.text == "Отмена":
        await cancel_action(message, state)
        return
    try:
        end_date = datetime.strptime(message.text.strip(), "%Y-%m-%d")
        await state.update_data(end_date=end_date)
        await state.set_state(BroadcastForm.waiting_for_text)
        await message.answer("Введите текст рассылки:")
    except:
        await message.answer("Неверный формат даты.")

@dp.message(BroadcastForm.waiting_for_text)
async def broadcast_text(message: types.Message, state: FSMContext):
    if message.text == "Отмена":
        await cancel_action(message, state)
        return
    data = await state.get_data()
    start_date = data["start_date"]
    end_date = data["end_date"]
    text = message.text.strip()
    start_dt = start_date.replace(tzinfo=timezone(timedelta(hours=3)))
    end_dt = end_date.replace(hour=23, minute=59, second=59, tzinfo=timezone(timedelta(hours=3)))
    res = supabase.table("tracks").select("user_id").gte("created_at", start_dt.isoformat()).lte("created_at", end_dt.isoformat()).execute()
    user_ids = list(set(t["user_id"] for t in res.data))
    if not user_ids:
        await message.answer("Нет пользователей за этот период.")
        await state.clear()
        return
    await message.answer(f"📢 Найдено {len(user_ids)} пользователей. Рассылаю...")
    sent = 0
    for uid in user_ids:
        try:
            await bot.send_message(uid, f"📢 {text}")
            sent += 1
            await asyncio.sleep(0.05)
        except:
            pass
    await message.answer(f"✅ Отправлено {sent} сообщений.")
    await state.clear()

# === ВЕБ-СЕРВЕР ===
async def handle_web(request):
    return web.Response(text="Bot is running")

async def start_web():
    app = web.Application()
    app.router.add_get("/", handle_web)
    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, "0.0.0.0", 8000)
    await site.start()
    print("Веб-сервер запущен")

async def run_bot():
    await bot.delete_webhook(drop_pending_updates=True)
    await dp.start_polling(bot)

async def main():
    await asyncio.gather(start_web(), run_bot())

if __name__ == "__main__":
    asyncio.run(main())
